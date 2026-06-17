import base64
import html
import json
import re
from pathlib import Path

import openpyxl


ROOT = Path(__file__).parent
SOURCE = ROOT / "ELS 2027 REPORTS_CURRICULUM.xlsm"
OUTPUT = ROOT / "early-learning-report-resource.html"
LOGO = ROOT / "assets" / "inspired-logo-wordmark.png"


def clean(value):
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    return re.sub(r"\n{3,}", "\n\n", text)


def normalize_area(value):
    value = clean(value).rstrip(":")
    for prefix, canonical in (
        ("Literacy", "Literacy"),
        ("Mathematics", "Mathematics"),
        ("Understanding the World", "Understanding the World"),
    ):
        if value.startswith(prefix):
            return canonical
    return value


def extract():
    workbook = openpyxl.load_workbook(
        SOURCE, data_only=True, read_only=False, keep_vba=True
    )
    records = []
    checkpoints = []
    areas = set()
    stage_meta = []
    groups = [
        (1, "Benchmarking"),
        (6, "Term 2"),
        (11, "Term 4"),
    ]

    for sheet in workbook.worksheets:
        if not sheet.title.startswith("STAGE"):
            continue

        stage = int(re.search(r"\d+", sheet.title).group())
        stage_entries = []
        for start_col, period in groups:
            age = clean(sheet.cell(1, start_col).value)
            source_period = clean(sheet.cell(2, start_col).value)
            checkpoint_id = f"s{stage}-{period.lower().replace(' ', '-')}"
            checkpoint = {
                "id": checkpoint_id,
                "stage": stage,
                "age": age,
                "period": period,
                "sourcePeriod": source_period,
            }
            checkpoints.append(checkpoint)
            stage_entries.append(checkpoint)

            current_area = ""
            current_overview = ""
            current_area_tips = ""
            for row in range(4, sheet.max_row + 1):
                category = clean(sheet.cell(row, start_col).value)
                statement = clean(sheet.cell(row, start_col + 1).value)
                guidance = clean(sheet.cell(row, start_col + 2).value)
                tips = clean(sheet.cell(row, start_col + 3).value)

                if not any((category, statement, guidance, tips)):
                    continue

                if category and not statement and not guidance:
                    next_category = clean(sheet.cell(row + 1, start_col).value)
                    next_statement = clean(sheet.cell(row + 1, start_col + 1).value)
                    next_guidance = clean(sheet.cell(row + 1, start_col + 2).value)
                    if next_category and not next_statement and not next_guidance:
                        current_area = normalize_area(category)
                        areas.add(current_area)
                        current_overview = clean(sheet.cell(row + 1, start_col).value)
                        current_area_tips = clean(
                            sheet.cell(row + 1, start_col + 3).value
                        )
                        continue

                if statement and current_area:
                    records.append(
                        {
                            "id": f"{checkpoint_id}-r{row}",
                            "checkpointId": checkpoint_id,
                            "stage": stage,
                            "age": age,
                            "period": period,
                            "area": current_area,
                            "category": category,
                            "statement": statement,
                            "guidance": guidance,
                            "tips": tips,
                            "overview": current_overview,
                            "areaTips": current_area_tips,
                            "sourceRow": row,
                        }
                    )

        stage_meta.append({"stage": stage, "checkpoints": stage_entries})

    notes = workbook["Notes"]
    cycle = []
    for row in range(3, 7):
        cycle.append(
            {
                "period": clean(notes.cell(row, 1).value),
                "ages": [
                    clean(notes.cell(row, col).value) for col in range(2, 7)
                ],
            }
        )

    return {
        "records": records,
        "checkpoints": checkpoints,
        "areas": sorted(areas),
        "stages": stage_meta,
        "cycle": cycle,
        "source": SOURCE.name,
    }


def render(data):
    payload = json.dumps(data, ensure_ascii=False).replace("</", "<\\/")
    logo_data = base64.b64encode(LOGO.read_bytes()).decode("ascii")
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>ELS 2025 Report Preparation Resource</title>
  <style>
    :root {{
      --ink:#143256; --muted:#81818c; --paper:#f5f5f6; --card:#ffffff;
      --teal:#214b76; --teal-dark:#122147; --coral:#e88b45; --gold:#e88b45;
      --line:#e1e3e7; --shadow:0 14px 34px rgba(18,33,71,.10);
    }}
    *{{box-sizing:border-box}}
    body{{margin:0;background:var(--paper);color:var(--ink);font-family:"Open Sans",Helvetica,Arial,sans-serif}}
    button,input,select{{font:inherit}}
    .top{{background:linear-gradient(115deg,rgba(18,33,71,.98),rgba(20,50,86,.92) 58%,rgba(33,75,118,.84));color:white;position:relative;overflow:hidden}}
    .top:after{{content:"";position:absolute;inset:0;background:linear-gradient(90deg,transparent 60%,rgba(232,139,69,.16));pointer-events:none}}
    .top-inner{{max-width:1220px;margin:auto;padding:25px 24px 92px;position:relative;z-index:1}}
    .brand-row{{display:flex;justify-content:space-between;align-items:center;border-bottom:1px solid rgba(255,255,255,.22);padding-bottom:20px;margin-bottom:58px}}
    .brand{{display:inline-flex;align-items:center;background:#fff;padding:9px 13px;border-radius:2px;box-shadow:0 4px 16px rgba(0,0,0,.08)}}
    .brand img{{display:block;width:170px;height:auto}}
    .brand-note{{font-family:Roboto,Helvetica,Arial,sans-serif;text-transform:uppercase;letter-spacing:.11em;font-size:.72rem;opacity:.82}}
    .eyebrow{{text-transform:uppercase;letter-spacing:.16em;font-size:.74rem;font-weight:600;color:#f1a76e}}
    h1{{font-family:Roboto,Helvetica,Arial,sans-serif;font-size:clamp(2.3rem,5vw,4.2rem);line-height:1.05;letter-spacing:-.045em;margin:14px 0 16px;max-width:820px;font-weight:300}}
    .intro{{max-width:720px;font-size:1.05rem;line-height:1.7;opacity:.88;margin:0;font-weight:300}}
    .main{{max-width:1220px;margin:-54px auto 70px;padding:0 24px;position:relative;z-index:2}}
    .control-panel{{background:var(--card);box-shadow:var(--shadow);border-top:4px solid var(--coral);border-radius:3px;padding:24px}}
    .controls{{display:grid;grid-template-columns:1.25fr 1fr 1fr 1.7fr;gap:12px}}
    label{{display:block;font-weight:600;font-size:.71rem;text-transform:uppercase;letter-spacing:.12em;color:var(--muted);margin:0 0 7px}}
    select,input{{width:100%;border:1px solid var(--line);border-radius:3px;padding:12px 13px;background:white;color:var(--ink);outline:none}}
    select:focus,input:focus{{border-color:var(--teal);box-shadow:0 0 0 3px rgba(22,125,115,.12)}}
    .quick{{display:flex;gap:8px;flex-wrap:wrap;margin-top:18px}}
    .quick button{{border:1px solid #d9dee5;background:#f7f8fa;color:var(--teal-dark);padding:8px 14px;border-radius:999px;font-weight:600;cursor:pointer}}
    .quick button:hover{{background:var(--teal-dark);color:white;border-color:var(--teal-dark)}}
    .summary{{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin:24px 0}}
    .metric{{background:var(--card);border:1px solid var(--line);border-radius:3px;padding:19px;border-bottom:3px solid #d8dee7}}
    .metric strong{{font-family:Roboto,Helvetica,Arial,sans-serif;font-size:2rem;font-weight:300;display:block;color:var(--teal-dark)}}
    .metric span{{color:var(--muted);font-size:.82rem;font-weight:500}}
    .section-head{{display:flex;align-items:end;justify-content:space-between;gap:20px;margin:36px 0 16px}}
    h2{{font-family:Roboto,Helvetica,Arial,sans-serif;font-weight:300;letter-spacing:-.03em;font-size:2rem;margin:0}}
    .result-text{{color:var(--muted);margin:0}}
    .area-overview{{display:none;background:#f7f8fa;border-left:4px solid var(--coral);border-radius:0;padding:20px;margin:0 0 18px}}
    .area-overview.show{{display:block}}
    .area-overview h3{{margin:0 0 8px;font-family:Roboto,Helvetica,Arial,sans-serif;font-weight:400;font-size:1.35rem}}
    .area-overview p{{margin:0;line-height:1.65}}
    .grid{{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:16px}}
    .card{{background:var(--card);border:1px solid var(--line);border-radius:3px;padding:21px;display:flex;flex-direction:column;gap:13px;transition:.18s;min-width:0}}
    .card:hover{{transform:translateY(-2px);box-shadow:0 14px 30px rgba(18,33,71,.10);border-color:#cbd2dc}}
    .card-top{{display:flex;justify-content:space-between;gap:10px;align-items:start}}
    .tag{{background:#e9eef4;color:var(--teal-dark);padding:6px 9px;border-radius:2px;font-size:.7rem;font-weight:600;letter-spacing:.07em;text-transform:uppercase}}
    .age{{font-size:.76rem;color:var(--muted);font-weight:750;text-align:right}}
    .card h3{{font-family:Roboto,Helvetica,Arial,sans-serif;font-weight:400;font-size:1.3rem;line-height:1.25;margin:0}}
    .statement{{font-size:1rem;line-height:1.62;margin:0}}
    details{{border-top:1px solid var(--line);padding-top:12px}}
    summary{{cursor:pointer;color:var(--teal-dark);font-weight:850;list-style:none}}
    summary::-webkit-details-marker{{display:none}}
    summary:after{{content:"+";float:right;font-size:1.2rem}}
    details[open] summary:after{{content:"−"}}
    .detail-block{{margin-top:13px;padding:13px 14px;border-radius:2px;background:#f5f6f8;line-height:1.58;font-size:.91rem;white-space:pre-line}}
    .detail-block b{{display:block;margin-bottom:5px;color:var(--teal-dark)}}
    .empty{{grid-column:1/-1;background:white;border:1px dashed #b8ccc4;border-radius:18px;text-align:center;padding:50px 20px;color:var(--muted)}}
    .guide{{margin-top:42px;background:var(--teal-dark);color:white;border-radius:3px;padding:29px;border-top:4px solid var(--coral)}}
    .guide h2{{margin-bottom:8px}}
    .guide p{{color:#d7e6e1;max-width:760px;line-height:1.55}}
    .cycle{{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-top:20px}}
    .cycle-card{{background:rgba(255,255,255,.07);border:1px solid rgba(255,255,255,.16);border-radius:2px;padding:14px}}
    .cycle-card strong{{display:block;color:#f1a76e;margin-bottom:8px;font-weight:600}}
    .cycle-card span{{display:block;font-size:.78rem;line-height:1.65;color:#e7f1ee}}
    footer{{max-width:1220px;margin:0 auto;padding:0 24px 35px;color:var(--muted);font-size:.78rem}}
    @media(max-width:900px){{.controls{{grid-template-columns:1fr 1fr}}.summary,.cycle{{grid-template-columns:1fr 1fr}}}}
    @media(max-width:650px){{.top-inner{{padding:22px 18px 76px}}.brand-row{{margin-bottom:42px}}.brand-note{{display:none}}.main{{padding:0 14px}}.controls,.summary,.grid,.cycle{{grid-template-columns:1fr}}.control-panel{{padding:16px}}.section-head{{display:block}}.result-text{{margin-top:7px}}}}
  </style>
</head>
<body>
  <header class="top">
    <div class="top-inner">
      <div class="brand-row"><div class="brand"><img src="data:image/png;base64,{logo_data}" alt="Inspired"></div><div class="brand-note">Early Learning School · Teacher Resource</div></div>
      <div class="eyebrow">2027 Report Preparation</div>
      <h1>Prepare meaningful 2027 reports with confidence.</h1>
      <p class="intro">Find the right developmental indicators, recognise what achievement looks like, and gather authentic evidence from everyday learning.</p>
    </div>
  </header>
  <main class="main">
    <section class="control-panel" aria-label="Dashboard filters">
      <div class="controls">
        <div><label for="stage">Stage</label><select id="stage"></select></div>
        <div><label for="period">Report checkpoint</label><select id="period"></select></div>
        <div><label for="area">Curriculum area</label><select id="area"></select></div>
        <div><label for="search">Search the resource</label><input id="search" type="search" placeholder="Try: turn-taking, counting, confidence…"></div>
      </div>
      <div class="quick" id="quick"></div>
    </section>

    <section class="summary" aria-label="Current selection summary">
      <div class="metric"><strong id="metricStage">—</strong><span>Selected stage</span></div>
      <div class="metric"><strong id="metricAge">—</strong><span>Developmental age</span></div>
      <div class="metric"><strong id="metricCount">—</strong><span>Indicators shown</span></div>
      <div class="metric"><strong id="metricAreas">—</strong><span>Curriculum areas</span></div>
    </section>

    <div class="section-head">
      <h2>Reporting indicators</h2>
      <p class="result-text" id="resultText"></p>
    </div>
    <aside class="area-overview" id="areaOverview"></aside>
    <section class="grid" id="cards"></section>

    <section class="guide">
      <h2>Reporting cycle at a glance</h2>
      <p>Use the age checkpoints as a guide, then rely on repeated, authentic observations across ordinary routines and play. The dashboard preserves the workbook’s guidance and practical teacher tips.</p>
      <div class="cycle" id="cycle"></div>
    </section>
  </main>
  <footer>Source: {html.escape(data["source"])} · Content is displayed as supplied in the curriculum workbook.</footer>
  <script>const DATA={payload};</script>
  <script>
    const $ = id => document.getElementById(id);
    const state = {{stage:"1", period:"Benchmarking", area:"All", search:""}};
    const esc = s => String(s||"").replace(/[&<>"']/g,m=>({{"&":"&amp;","<":"&lt;",">":"&gt;",'"':"&quot;","'":"&#039;"}}[m]));
    const stageSel=$("stage"),periodSel=$("period"),areaSel=$("area"),search=$("search");
    stageSel.innerHTML=DATA.stages.map(s=>`<option value="${{s.stage}}">Stage ${{s.stage}}</option>`).join("");
    periodSel.innerHTML=["Benchmarking","Term 2","Term 4"].map(x=>`<option>${{x}}</option>`).join("");
    areaSel.innerHTML=`<option>All</option>`+DATA.areas.map(x=>`<option>${{esc(x)}}</option>`).join("");
    $("quick").innerHTML=DATA.areas.map(x=>`<button data-area="${{esc(x)}}">${{esc(x)}}</button>`).join("");
    $("quick").onclick=e=>{{if(e.target.dataset.area){{areaSel.value=e.target.dataset.area;state.area=e.target.dataset.area;render()}}}};
    [stageSel,periodSel,areaSel].forEach(el=>el.onchange=()=>{{state.stage=stageSel.value;state.period=periodSel.value;state.area=areaSel.value;render()}});
    search.oninput=()=>{{state.search=search.value.trim().toLowerCase();render()}};

    function currentCheckpoint(){{
      return DATA.checkpoints.find(x=>String(x.stage)===state.stage&&x.period===state.period);
    }}
    function render(){{
      const q=state.search;
      const filtered=DATA.records.filter(r=>String(r.stage)===state.stage&&r.period===state.period&&(state.area==="All"||r.area===state.area)&&(!q||[r.area,r.category,r.statement,r.guidance,r.tips,r.overview].join(" ").toLowerCase().includes(q)));
      const checkpoint=currentCheckpoint();
      const areaCount=new Set(filtered.map(x=>x.area)).size;
      $("metricStage").textContent=`Stage ${{state.stage}}`;
      $("metricAge").textContent=(checkpoint?.age||"—").replace("Age ","");
      $("metricCount").textContent=filtered.length;
      $("metricAreas").textContent=areaCount;
      $("resultText").textContent=`${{filtered.length}} indicator${{filtered.length===1?"":"s"}} · ${{checkpoint?.age||""}} · ${{state.period}}`;
      const selectedArea=state.area!=="All" ? filtered[0] || DATA.records.find(r=>String(r.stage)===state.stage&&r.period===state.period&&r.area===state.area) : null;
      const overview=$("areaOverview");
      if(selectedArea?.overview){{
        overview.className="area-overview show";
        overview.innerHTML=`<h3>${{esc(state.area)}} overview</h3><p>${{esc(selectedArea.overview)}}</p>`;
      }} else {{overview.className="area-overview"; overview.innerHTML=""}}
      $("cards").innerHTML=filtered.length?filtered.map(r=>`
        <article class="card">
          <div class="card-top"><span class="tag">${{esc(r.area)}}</span><span class="age">Stage ${{r.stage}} · ${{esc(r.age)}}<br>${{esc(r.period)}}</span></div>
          <h3>${{esc(r.category)}}</h3>
          <p class="statement">${{esc(r.statement)}}</p>
          <details><summary>What achievement may look like</summary><div class="detail-block">${{esc(r.guidance)}}</div></details>
          <details><summary>Practical evidence ideas</summary><div class="detail-block">${{esc(r.tips||r.areaTips||"Observe during authentic routines and child-led play, recording repeated examples over time.")}}</div></details>
        </article>`).join(""):`<div class="empty"><h3>No matching indicators</h3><p>Try clearing the search or choosing a different area.</p></div>`;
    }}
    $("cycle").innerHTML=DATA.cycle.map(x=>`<div class="cycle-card"><strong>${{esc(x.period)}}</strong><span>${{x.ages.map((a,i)=>`Stage ${{i+1}}: ${{esc(a)}}`).join("<br>")}}</span></div>`).join("");
    render();
  </script>
</body>
</html>"""


if __name__ == "__main__":
    extracted = extract()
    OUTPUT.write_text(render(extracted), encoding="utf-8")
    summary = {
        "output": str(OUTPUT),
        "records": len(extracted["records"]),
        "checkpoints": len(extracted["checkpoints"]),
        "areas": extracted["areas"],
        "stages": len(extracted["stages"]),
    }
    print(json.dumps(summary, indent=2, ensure_ascii=False))
